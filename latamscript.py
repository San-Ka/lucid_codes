#excel file name
Input_File_name = "/Users/skaur2/Desktop/latam.xlsx"
Output_file_name = "/Users/skaur2/Desktop/output1.xlsx"
country_row_name = 'country'
client_row_name = 'client'
buyer_account_row_name = 'buyer_account'
buyer_account_id_row_name = 'buyer_account_id'
buyer_bu_id_row_name = 'buyer_bu_id'
drops_row_name = 'drops'
total_entrants_row_name = 'total_entrants'
completes_row_name = 'completes'
survey_number_row_name = 'surveynumber'
project_manager_row_name = 'projectmanager'
project_manager_email_row_name = 'pm_email'
drop_rate_row_name = 'drop_rate'
system_conversion_row_name =  'system_conversion'
########################################################################################################################
import xlrd
import collections
import openpyxl
from collections import defaultdict
#open excel file
workbook = xlrd.open_workbook(Input_File_name)
worksheet = workbook.sheet_by_index(0)
first_row = []
for col in range(worksheet.ncols):
   first_row.append( worksheet.cell_value(0,col) )
#get excel data in list
data =[]
for row in range(1, worksheet.nrows):
   elm = {}
   for col in range(worksheet.ncols):
       elm[first_row[col]]=worksheet.cell_value(row,col)
   data.append(elm)
surveyNo_Pm_dict = {}
surveyNo_droprate_dict = {}
surveyNo_systemconversion_dict = {}
pmemail_pm_dict = {}
pmemail_email_dict = {}
#
pm_email_country_dict = {}
pm_email_client_dict = {}
pm_email_buyer_account_dict = {}
pm_email_buyer_account_id_dict = {}
pm_email_buyer_bu_id_dict = {}
pm_email_drops_dict = {}
pm_email_total_entrants_dict = {}
pm_email_completes_dict = {}
for i in data:
 survey_no = int(i[survey_number_row_name])
 pm = i[project_manager_row_name]
 pm_email = i[project_manager_email_row_name]
 pm_pmemail_concat = pm+pm_email
 droprate = i[drop_rate_row_name] * 100 
 systemconversion = i[system_conversion_row_name] * 100
 #
 country = i[country_row_name]
 client = i[client_row_name]
 buyer_account = i[buyer_account_row_name]
 buyer_account_id = i[buyer_account_id_row_name]
 buyer_bu_id = i[buyer_bu_id_row_name]
 drops = i[drops_row_name]
 total_entrants = i[total_entrants_row_name]
 completes = i[completes_row_name]
 if isinstance(droprate, str) != True:
   droprate = int(droprate)
 if isinstance(systemconversion, str) !=True:
     systemconversion = int(systemconversion) 
 surveyNo_Pm_dict[survey_no] = pm+pm_email
 surveyNo_droprate_dict[survey_no] = droprate 
 surveyNo_systemconversion_dict[survey_no] = systemconversion
 pmemail_pm_dict[pm_pmemail_concat] = pm
 pmemail_email_dict[pm_pmemail_concat] = pm_email
 #
 pm_email_country_dict[pm_pmemail_concat] = country 
 pm_email_client_dict[pm_pmemail_concat] = client
 pm_email_buyer_account_dict[pm_pmemail_concat] = buyer_account
 pm_email_buyer_account_id_dict[pm_pmemail_concat] = buyer_account_id
 pm_email_buyer_bu_id_dict[pm_pmemail_concat] = buyer_bu_id
 pm_email_drops_dict[pm_pmemail_concat] = drops
 pm_email_total_entrants_dict[pm_pmemail_concat] = total_entrants
 pm_email_completes_dict[pm_pmemail_concat] = completes
#reverse dictionary if it has same value append it
Pm_surveyNo_dict = defaultdict(list)
for key,value in surveyNo_Pm_dict.items():
 Pm_surveyNo_dict[value].append(key)
new_Pm_surveyNo_dict = {}
for key, value in Pm_surveyNo_dict.items():
 new_value = []
 for i in value:
   new_value.append(surveyNo_droprate_dict[i])
   new_value.append(surveyNo_systemconversion_dict[i])
# country_survye_no_dict = defaultdict(list)
# for key,value in survey_no_country_dict.items():
#  country_survye_no_dict[value].append(key)
# # #Insert data in excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1, value='Survey Number')
sheet.cell(row=1, column=2, value ='Country')
sheet.cell(row=1, column=3, value ='Client Name')
sheet.cell(row=1, column=4, value = 'Buyer Account')
sheet.cell(row=1, column=5, value = 'Buyer Account ID')
sheet.cell(row=1, column=6, value = 'Buyer Account BU ID')
sheet.cell(row=1, column=7, value='Project Manager')
sheet.cell(row=1, column=8, value='Project Manager Email')
sheet.cell(row=1, column=9, value = 'Drops')
sheet.cell(row=1, column=10, value='Total Entrants')
sheet.cell(row=1, column=11, value='Completes' )
sheet.cell(row=1, column=12, value='Drop Rate')
sheet.cell(row=1, column=13, value='System Conversion')
row = 2
for key,values in Pm_surveyNo_dict.items():
   sheet.cell(row=row, column=2,value=pm_email_country_dict[key])
   sheet.cell(row=row, column=3,value=pm_email_client_dict[key])
   sheet.cell(row=row, column=4,value=pm_email_buyer_account_dict[key])
   sheet.cell(row=row, column=5,value=pm_email_buyer_account_id_dict[key])
   sheet.cell(row=row, column=6, value=pm_email_buyer_bu_id_dict[key])
   sheet.cell(row=row, column=7, value=pmemail_pm_dict[key])
   sheet.cell(row=row, column=8, value=pmemail_email_dict[key])
   sheet.cell(row=row, column=9, value=pm_email_drops_dict[key])
   sheet.cell(row=row, column=10, value=pm_email_total_entrants_dict[key])
   sheet.cell(row=row, column=11, value=pm_email_completes_dict[key])
   column = 2
   survey_no_sentence = []
   droprate_sentence = []
   systemconversion_sentence = []
   for survey_no in values:
       droprate = surveyNo_droprate_dict[survey_no]
       survey_no_sentence.append(str(survey_no))
       droprate_sentence.append(str(droprate) + '%') 
       systemconversion = surveyNo_systemconversion_dict[survey_no]
       systemconversion_sentence.append(str(systemconversion) + '%')
   survey_no_collection = '<br>'.join(survey_no_sentence)
   droprate_collection = '<br>'.join(droprate_sentence)
   systemconversion_collection = '<br>'.join(systemconversion_sentence)
   sheet.cell(row=row, column=1, value=str(survey_no_collection))
   sheet.cell(row=row, column=12, value=str(droprate_collection))
   sheet.cell(row=row, column=13, value=str(systemconversion_collection))
 
   row += 1
#will save the excel file
workbook.save(filename=Output_file_name)